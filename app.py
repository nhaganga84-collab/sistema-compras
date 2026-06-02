from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import os, io, json

app = Flask(__name__)
app.config['SECRET_KEY'] = 'Sistema de compas-mz-2026'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///compras.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}

# ─────────────────────────── MODELOS ───────────────────────────

class Fornecedor(db.Model):
    __tablename__ = 'suppliers'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(150), nullable=False)
    nuit = db.Column(db.String(20))
    telefone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    endereco = db.Column(db.Text)
    contacto = db.Column(db.String(100))
    estado = db.Column(db.String(10), default='Ativo')
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)
    produtos = db.relationship('Produto', backref='fornecedor_principal', lazy=True)
    compras = db.relationship('Compra', backref='fornecedor', lazy=True)
    pedidos = db.relationship('PedidoCompra', backref='fornecedor', lazy=True)
    contas = db.relationship('ContaPagar', backref='fornecedor', lazy=True)

class Produto(db.Model):
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20), unique=True, nullable=False)
    nome = db.Column(db.String(150), nullable=False)
    categoria = db.Column(db.String(80))
    preco_compra = db.Column(db.Float, default=0)
    preco_venda = db.Column(db.Float, default=0)
    stock = db.Column(db.Integer, default=0)
    stock_minimo = db.Column(db.Integer, default=5)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'))
    itens = db.relationship('ItemCompra', backref='produto', lazy=True)
    itens_pedido = db.relationship('ItemPedido', backref='produto', lazy=True)

class PedidoCompra(db.Model):
    __tablename__ = 'purchase_orders'
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(10), unique=True)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'))
    data_pedido = db.Column(db.DateTime, default=datetime.utcnow)
    estado = db.Column(db.String(20), default='Pendente')
    total = db.Column(db.Float, default=0)
    observacoes = db.Column(db.Text)
    itens = db.relationship('ItemPedido', backref='pedido', lazy=True, cascade='all, delete-orphan')

class ItemPedido(db.Model):
    __tablename__ = 'order_items'
    id = db.Column(db.Integer, primary_key=True)
    pedido_id = db.Column(db.Integer, db.ForeignKey('purchase_orders.id'))
    produto_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    quantidade = db.Column(db.Integer, default=1)
    preco_unitario = db.Column(db.Float, default=0)
    subtotal = db.Column(db.Float, default=0)

class Compra(db.Model):
    __tablename__ = 'purchases'
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(10), unique=True)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'))
    data_compra = db.Column(db.DateTime, default=datetime.utcnow)
    total = db.Column(db.Float, default=0)
    metodo_pagamento = db.Column(db.String(30))
    estado_pagamento = db.Column(db.String(20), default='Pendente')
    observacoes = db.Column(db.Text)
    itens = db.relationship('ItemCompra', backref='compra', lazy=True, cascade='all, delete-orphan')

class ItemCompra(db.Model):
    __tablename__ = 'purchase_items'
    id = db.Column(db.Integer, primary_key=True)
    compra_id = db.Column(db.Integer, db.ForeignKey('purchases.id'))
    produto_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    quantidade = db.Column(db.Integer, default=1)
    preco_unitario = db.Column(db.Float, default=0)
    subtotal = db.Column(db.Float, default=0)

class ContaPagar(db.Model):
    __tablename__ = 'accounts_payable'
    id = db.Column(db.Integer, primary_key=True)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'))
    compra_id = db.Column(db.Integer, db.ForeignKey('purchases.id'), nullable=True)
    valor = db.Column(db.Float, default=0)
    vencimento = db.Column(db.Date)
    estado = db.Column(db.String(15), default='Pendente')
    descricao = db.Column(db.Text)
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)
    compra = db.relationship('Compra', backref='conta', lazy=True)

# ─────────────────────────── HELPERS ───────────────────────────

def gerar_numero(modelo, prefixo, campo='numero'):
    ultimo = db.session.query(modelo).order_by(modelo.id.desc()).first()
    num = (ultimo.id + 1) if ultimo else 1
    return f"{prefixo}{num:04d}"

def atualizar_estado_contas():
    hoje = date.today()
    contas = ContaPagar.query.filter_by(estado='Pendente').all()
    for c in contas:
        if c.vencimento and c.vencimento < hoje:
            c.estado = 'Atrasado'
    db.session.commit()

# ─────────────────────────── DASHBOARD ───────────────────────────

@app.route('/')
def dashboard():
    atualizar_estado_contas()
    hoje = datetime.utcnow()
    compras_mes = Compra.query.filter(
        db.extract('month', Compra.data_compra) == hoje.month,
        db.extract('year', Compra.data_compra) == hoje.year
    ).all()
    valor_mes = sum(c.total for c in compras_mes)
    stock_baixo = Produto.query.filter(Produto.stock <= Produto.stock_minimo).all()
    contas_pendentes = ContaPagar.query.filter(ContaPagar.estado.in_(['Pendente', 'Atrasado'])).all()
    valor_pendente = sum(c.valor for c in contas_pendentes)
    contas_atrasadas = ContaPagar.query.filter_by(estado='Atrasado').count()

    # Dados para gráfico (últimos 6 meses)
    from dateutil.relativedelta import relativedelta
    meses_dados = []
    for i in range(5, -1, -1):
        d = hoje - relativedelta(months=i)
        total = db.session.query(db.func.sum(Compra.total)).filter(
            db.extract('month', Compra.data_compra) == d.month,
            db.extract('year', Compra.data_compra) == d.year
        ).scalar() or 0
        meses_dados.append({'mes': d.strftime('%b/%y'), 'total': round(total, 2)})

    return render_template('dashboard.html',
        total_fornecedores=Fornecedor.query.filter_by(estado='Ativo').count(),
        compras_mes=len(compras_mes),
        valor_mes=valor_mes,
        stock_baixo=stock_baixo,
        contas_pendentes=len(contas_pendentes),
        valor_pendente=valor_pendente,
        contas_atrasadas=contas_atrasadas,
        meses_dados=json.dumps(meses_dados)
    )

# ─────────────────────────── FORNECEDORES ───────────────────────────

@app.route('/fornecedores')
def fornecedores():
    q = request.args.get('q', '')
    query = Fornecedor.query
    if q:
        query = query.filter(db.or_(Fornecedor.nome.ilike(f'%{q}%'), Fornecedor.nuit.ilike(f'%{q}%')))
    lista = query.order_by(Fornecedor.nome).all()
    return render_template('fornecedores.html', lista=lista, q=q)

@app.route('/fornecedores/novo', methods=['GET', 'POST'])
def novo_fornecedor():
    if request.method == 'POST':
        f = Fornecedor(
            nome=request.form['nome'],
            nuit=request.form.get('nuit', ''),
            telefone=request.form.get('telefone', ''),
            email=request.form.get('email', ''),
            endereco=request.form.get('endereco', ''),
            contacto=request.form.get('contacto', ''),
            estado=request.form.get('estado', 'Ativo')
        )
        db.session.add(f)
        db.session.commit()
        flash('Fornecedor cadastrado com sucesso!', 'success')
        return redirect(url_for('fornecedores'))
    return render_template('fornecedor_form.html', fornecedor=None)

@app.route('/fornecedores/<int:id>/editar', methods=['GET', 'POST'])
def editar_fornecedor(id):
    f = Fornecedor.query.get_or_404(id)
    if request.method == 'POST':
        f.nome = request.form['nome']
        f.nuit = request.form.get('nuit', '')
        f.telefone = request.form.get('telefone', '')
        f.email = request.form.get('email', '')
        f.endereco = request.form.get('endereco', '')
        f.contacto = request.form.get('contacto', '')
        f.estado = request.form.get('estado', 'Ativo')
        db.session.commit()
        flash('Fornecedor atualizado!', 'success')
        return redirect(url_for('fornecedores'))
    return render_template('fornecedor_form.html', fornecedor=f)

@app.route('/fornecedores/<int:id>/remover', methods=['POST'])
def remover_fornecedor(id):
    f = Fornecedor.query.get_or_404(id)
    f.estado = 'Inativo'
    db.session.commit()
    flash('Fornecedor desativado.', 'info')
    return redirect(url_for('fornecedores'))

# ─────────────────────────── PRODUTOS ───────────────────────────

@app.route('/produtos')
def produtos():
    q = request.args.get('q', '')
    query = Produto.query
    if q:
        query = query.filter(db.or_(Produto.nome.ilike(f'%{q}%'), Produto.codigo.ilike(f'%{q}%'), Produto.categoria.ilike(f'%{q}%')))
    lista = query.order_by(Produto.nome).all()
    return render_template('produtos.html', lista=lista, q=q)

@app.route('/produtos/novo', methods=['GET', 'POST'])
def novo_produto():
    fornecedores = Fornecedor.query.filter_by(estado='Ativo').order_by(Fornecedor.nome).all()
    if request.method == 'POST':
        p = Produto(
            codigo=request.form['codigo'],
            nome=request.form['nome'],
            categoria=request.form.get('categoria', ''),
            preco_compra=float(request.form.get('preco_compra', 0)),
            preco_venda=float(request.form.get('preco_venda', 0)),
            stock=int(request.form.get('stock', 0)),
            stock_minimo=int(request.form.get('stock_minimo', 5)),
            fornecedor_id=request.form.get('fornecedor_id') or None
        )
        db.session.add(p)
        db.session.commit()
        flash('Produto cadastrado!', 'success')
        return redirect(url_for('produtos'))
    return render_template('produto_form.html', produto=None, fornecedores=fornecedores)

@app.route('/produtos/<int:id>/editar', methods=['GET', 'POST'])
def editar_produto(id):
    p = Produto.query.get_or_404(id)
    fornecedores = Fornecedor.query.filter_by(estado='Ativo').order_by(Fornecedor.nome).all()
    if request.method == 'POST':
        p.codigo = request.form['codigo']
        p.nome = request.form['nome']
        p.categoria = request.form.get('categoria', '')
        p.preco_compra = float(request.form.get('preco_compra', 0))
        p.preco_venda = float(request.form.get('preco_venda', 0))
        p.stock = int(request.form.get('stock', 0))
        p.stock_minimo = int(request.form.get('stock_minimo', 5))
        p.fornecedor_id = request.form.get('fornecedor_id') or None
        db.session.commit()
        flash('Produto atualizado!', 'success')
        return redirect(url_for('produtos'))
    return render_template('produto_form.html', produto=p, fornecedores=fornecedores)

# ─────────────────────────── PEDIDOS DE COMPRA ───────────────────────────

@app.route('/pedidos')
def pedidos():
    lista = PedidoCompra.query.order_by(PedidoCompra.id.desc()).all()
    return render_template('pedidos.html', lista=lista)

@app.route('/pedidos/novo', methods=['GET', 'POST'])
def novo_pedido():
    fornecedores = Fornecedor.query.filter_by(estado='Ativo').order_by(Fornecedor.nome).all()
    produtos_obj = Produto.query.order_by(Produto.nome).all()
    produtos = produtos_obj
    produtos_json = [{'id': p.id, 'nome': p.nome, 'codigo': p.codigo,
                      'preco_compra': p.preco_compra, 'stock': p.stock} for p in produtos_obj]
    if request.method == 'POST':
        pedido = PedidoCompra(
            numero=gerar_numero(PedidoCompra, 'PED'),
            fornecedor_id=request.form['fornecedor_id'],
            estado='Pendente',
            observacoes=request.form.get('observacoes', '')
        )
        db.session.add(pedido)
        db.session.flush()
        total = 0
        ids = request.form.getlist('produto_id[]')
        qtds = request.form.getlist('quantidade[]')
        precos = request.form.getlist('preco_unitario[]')
        for pid, qtd, preco in zip(ids, qtds, precos):
            if pid and qtd:
                sub = float(qtd) * float(preco)
                total += sub
                item = ItemPedido(pedido_id=pedido.id, produto_id=int(pid),
                                  quantidade=int(qtd), preco_unitario=float(preco), subtotal=sub)
                db.session.add(item)
        pedido.total = total
        db.session.commit()
        flash(f'Pedido {pedido.numero} criado!', 'success')
        return redirect(url_for('pedidos'))
    return render_template('pedido_form.html', fornecedores=fornecedores, produtos=produtos_json)

@app.route('/pedidos/<int:id>/estado', methods=['POST'])
def alterar_estado_pedido(id):
    p = PedidoCompra.query.get_or_404(id)
    p.estado = request.form['estado']
    db.session.commit()
    flash(f'Estado atualizado para {p.estado}', 'info')
    return redirect(url_for('pedidos'))

@app.route('/pedidos/<int:id>')
def ver_pedido(id):
    p = PedidoCompra.query.get_or_404(id)
    return render_template('pedido_detalhe.html', pedido=p)

# ─────────────────────────── COMPRAS ───────────────────────────

@app.route('/compras')
def compras():
    lista = Compra.query.order_by(Compra.id.desc()).all()
    return render_template('compras.html', lista=lista)

@app.route('/compras/nova', methods=['GET', 'POST'])
def nova_compra():
    fornecedores = Fornecedor.query.filter_by(estado='Ativo').order_by(Fornecedor.nome).all()
    produtos_obj = Produto.query.order_by(Produto.nome).all()
    produtos = produtos_obj
    produtos_json = [{'id': p.id, 'nome': p.nome, 'codigo': p.codigo,
                      'preco_compra': p.preco_compra, 'stock': p.stock} for p in produtos_obj]
    if request.method == 'POST':
        compra = Compra(
            numero=gerar_numero(Compra, 'CMP'),
            fornecedor_id=request.form['fornecedor_id'],
            metodo_pagamento=request.form.get('metodo_pagamento', ''),
            estado_pagamento=request.form.get('estado_pagamento', 'Pendente'),
            observacoes=request.form.get('observacoes', '')
        )
        db.session.add(compra)
        db.session.flush()
        total = 0
        ids = request.form.getlist('produto_id[]')
        qtds = request.form.getlist('quantidade[]')
        precos = request.form.getlist('preco_unitario[]')
        for pid, qtd, preco in zip(ids, qtds, precos):
            if pid and qtd:
                sub = float(qtd) * float(preco)
                total += sub
                item = ItemCompra(compra_id=compra.id, produto_id=int(pid),
                                  quantidade=int(qtd), preco_unitario=float(preco), subtotal=sub)
                db.session.add(item)
                # Atualizar stock
                prod = Produto.query.get(int(pid))
                if prod:
                    prod.stock += int(qtd)
                    prod.preco_compra = float(preco)
        compra.total = total
        # Criar conta a pagar se pendente
        if compra.estado_pagamento == 'Pendente':
            venc = request.form.get('vencimento')
            venc_date = datetime.strptime(venc, '%Y-%m-%d').date() if venc else date.today()
            conta = ContaPagar(fornecedor_id=compra.fornecedor_id, compra_id=compra.id,
                               valor=total, vencimento=venc_date, estado='Pendente',
                               descricao=f'Compra {compra.numero}')
            db.session.add(conta)
        db.session.commit()
        flash(f'Compra {compra.numero} registada! Stock atualizado.', 'success')
        return redirect(url_for('compras'))
    return render_template('compra_form.html', fornecedores=fornecedores, produtos=produtos_json)

@app.route('/compras/<int:id>')
def ver_compra(id):
    c = Compra.query.get_or_404(id)
    return render_template('compra_detalhe.html', compra=c)

# ─────────────────────────── CONTAS A PAGAR ───────────────────────────

@app.route('/contas')
def contas():
    atualizar_estado_contas()
    lista = ContaPagar.query.order_by(ContaPagar.vencimento).all()
    return render_template('contas.html', lista=lista)

@app.route('/contas/nova', methods=['GET', 'POST'])
def nova_conta():
    fornecedores = Fornecedor.query.filter_by(estado='Ativo').order_by(Fornecedor.nome).all()
    if request.method == 'POST':
        venc = request.form.get('vencimento')
        conta = ContaPagar(
            fornecedor_id=request.form['fornecedor_id'],
            valor=float(request.form['valor']),
            vencimento=datetime.strptime(venc, '%Y-%m-%d').date() if venc else date.today(),
            estado='Pendente',
            descricao=request.form.get('descricao', '')
        )
        db.session.add(conta)
        db.session.commit()
        flash('Conta registada!', 'success')
        return redirect(url_for('contas'))
    return render_template('conta_form.html', fornecedores=fornecedores)

@app.route('/contas/<int:id>/pagar', methods=['POST'])
def pagar_conta(id):
    c = ContaPagar.query.get_or_404(id)
    c.estado = 'Pago'
    if c.compra_id:
        comp = Compra.query.get(c.compra_id)
        if comp:
            comp.estado_pagamento = 'Pago'
    db.session.commit()
    flash('Conta marcada como paga!', 'success')
    return redirect(url_for('contas'))

# ─────────────────────────── RELATÓRIOS ───────────────────────────

@app.route('/relatorios')
def relatorios():
    return render_template('relatorios.html')

@app.route('/relatorios/compras')
def rel_compras():
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    query = Compra.query
    if from_date:
        query = query.filter(Compra.data_compra >= datetime.strptime(from_date, '%Y-%m-%d'))
    if to_date:
        query = query.filter(Compra.data_compra <= datetime.strptime(to_date, '%Y-%m-%d').replace(hour=23, minute=59))
    lista = query.order_by(Compra.data_compra.desc()).all()
    total = sum(c.total for c in lista)
    return render_template('rel_compras.html', lista=lista, total=total, from_date=from_date, to_date=to_date)

@app.route('/relatorios/fornecedores')
def rel_fornecedores():
    fornecedores = Fornecedor.query.filter_by(estado='Ativo').all()
    dados = []
    for f in fornecedores:
        total = sum(c.total for c in f.compras)
        ultima = max((c.data_compra for c in f.compras), default=None)
        dados.append({'fornecedor': f, 'total': total, 'ultima': ultima, 'num_compras': len(f.compras)})
    dados.sort(key=lambda x: x['total'], reverse=True)
    return render_template('rel_fornecedores.html', dados=dados)

@app.route('/relatorios/financeiro')
def rel_financeiro():
    ano = request.args.get('ano', datetime.utcnow().year, type=int)
    meses = []
    for m in range(1, 13):
        total = db.session.query(db.func.sum(Compra.total)).filter(
            db.extract('month', Compra.data_compra) == m,
            db.extract('year', Compra.data_compra) == ano
        ).scalar() or 0
        meses.append({'mes': m, 'nome': ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'][m-1], 'total': round(total, 2)})
    total_ano = sum(m['total'] for m in meses)
    return render_template('rel_financeiro.html', meses=meses, total_ano=total_ano, ano=ano)

# ─────────────────────────── EXPORTAÇÕES ───────────────────────────

@app.route('/exportar/compras/pdf')
def exportar_compras_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    compras = Compra.query.order_by(Compra.data_compra.desc()).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph('Relatório de Compras', styles['Title']))
    story.append(Paragraph(f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    data = [['Nº', 'Fornecedor', 'Data', 'Total (MT)', 'Pagamento', 'Estado']]
    for c in compras:
        data.append([c.numero, c.fornecedor.nome if c.fornecedor else '-',
                     c.data_compra.strftime('%d/%m/%Y'), f'{c.total:,.2f}',
                     c.metodo_pagamento, c.estado_pagamento])

    t = Table(data, colWidths=[2.5*cm, 7*cm, 3*cm, 3.5*cm, 3.5*cm, 3*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))
    total = sum(c.total for c in compras)
    story.append(Paragraph(f'<b>Total Geral: MT {total:,.2f}</b>', styles['Normal']))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='compras.pdf', mimetype='application/pdf')

@app.route('/exportar/compras/excel')
def exportar_compras_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    compras = Compra.query.order_by(Compra.data_compra.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Compras'
    headers = ['Número', 'Fornecedor', 'Data', 'Total (MT)', 'Método Pagamento', 'Estado Pagamento']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1e3a5f')
        cell.alignment = Alignment(horizontal='center')
    for c in compras:
        ws.append([c.numero, c.fornecedor.nome if c.fornecedor else '-',
                   c.data_compra.strftime('%d/%m/%Y'), c.total,
                   c.metodo_pagamento, c.estado_pagamento])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='compras.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar/financeiro/pdf')
def exportar_financeiro_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    ano = datetime.utcnow().year
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f'Relatório Financeiro - {ano}', styles['Title']))
    story.append(Spacer(1, 0.5*cm))
    data = [['Mês', 'Total Gasto (MT)']]
    nomes = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    total_ano = 0
    for m in range(1, 13):
        total = db.session.query(db.func.sum(Compra.total)).filter(
            db.extract('month', Compra.data_compra) == m,
            db.extract('year', Compra.data_compra) == ano
        ).scalar() or 0
        total_ano += total
        data.append([nomes[m-1], f'MT {total:,.2f}'])
    data.append(['TOTAL', f'MT {total_ano:,.2f}'])
    t = Table(data, colWidths=[8*cm, 8*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e8f0fe')),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='financeiro.pdf', mimetype='application/pdf')

# ─────────────────────────── API JSON ───────────────────────────

@app.route('/api/produto/<int:id>')
def api_produto(id):
    p = Produto.query.get_or_404(id)
    return jsonify({'id': p.id, 'nome': p.nome, 'preco_compra': p.preco_compra, 'stock': p.stock})

@app.route('/api/stats')
def api_stats():
    return jsonify({
        'fornecedores': Fornecedor.query.filter_by(estado='Ativo').count(),
        'produtos': Produto.query.count(),
        'compras': Compra.query.count(),
        'stock_baixo': Produto.query.filter(Produto.stock <= Produto.stock_minimo).count()
    })

# ─────────────────────────── INIT ───────────────────────────

def seed_data():
    if Fornecedor.query.count() == 0:
        fornecedores = [
            Fornecedor(nome='ABC Distribuidora Lda', nuit='400123456', telefone='+258 84 111 2222',
                       email='abc@distribuidora.co.mz', endereco='Av. Eduardo Mondlane, 123, Maputo',
                       contacto='João Silva', estado='Ativo'),
            Fornecedor(nome='Sul Comércio & Serviços', nuit='400654321', telefone='+258 86 333 4444',
                       email='sul@comercio.co.mz', endereco='Rua da Resistência, 45, Matola',
                       contacto='Maria Machava', estado='Ativo'),
            Fornecedor(nome='Moçambique Food Supply', nuit='400987654', telefone='+258 82 555 6666',
                       email='info@mzfood.co.mz', endereco='Av. Julius Nyerere, 78, Maputo',
                       contacto='Pedro Nhantumbo', estado='Ativo'),
        ]
        db.session.add_all(fornecedores)
        db.session.flush()

        produtos = [
            Produto(codigo='PRD001', nome='Arroz 50kg', categoria='Alimentação', preco_compra=1200, preco_venda=1500, stock=80, stock_minimo=10, fornecedor_id=fornecedores[0].id),
            Produto(codigo='PRD002', nome='Açúcar 25kg', categoria='Alimentação', preco_compra=650, preco_venda=850, stock=45, stock_minimo=15, fornecedor_id=fornecedores[0].id),
            Produto(codigo='PRD003', nome='Óleo 20L', categoria='Alimentação', preco_compra=900, preco_venda=1100, stock=3, stock_minimo=10, fornecedor_id=fornecedores[1].id),
            Produto(codigo='PRD004', nome='Farinha 25kg', categoria='Alimentação', preco_compra=550, preco_venda=720, stock=2, stock_minimo=8, fornecedor_id=fornecedores[2].id),
            Produto(codigo='PRD005', nome='Sal 5kg', categoria='Temperos', preco_compra=120, preco_venda=180, stock=60, stock_minimo=20, fornecedor_id=fornecedores[1].id),
        ]
        db.session.add_all(produtos)
        db.session.commit()

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        seed_data()
    app.run(debug=True, port=5000)
